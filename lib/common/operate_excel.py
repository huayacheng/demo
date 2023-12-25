# -*- coding: utf-8 -*-
# operate_excel.py
import xlrd
import openpyxl
import re
from fuzzywuzzy import fuzz
class OperateExcel:

    # 构造函数初始化的时候会加载成员变量和类方法
    def __init__(self, file_path=None, sheet_name=None,rows_id=None):
        # 如果传入了excel文件路径则使用该路径，否则使用默认的excel文件路径
        # 如果传入sheet页的值则使用传入的值，否则默认是第一sheet页
        if file_path:
            self.file_path = file_path
            self.sheet_name = sheet_name
            self.rows_id = rows_id
        else:
            self.file_path = None
            self.sheet_id = 0
            self.rows_id = 0
        self.table = self.get_sheet(sheet_name)
        print(self.table)

    # 获取excel的所有sheet名称
    def get_sheet_name(self):
        # 打开excel文件读取数据
        if self.file_path != None:
            data = xlrd.open_workbook((self.file_path))
            print(data)
            sheet_name = []
            for sheet in data.sheet_names():
                print(sheet)
                sheet_name.append(sheet)
            print(sheet_name)
            return sheet_name

    # 获取excel的sheet页
    def get_sheet(self,sheet_name):
        # 打开excel文件读取数据
        if self.file_path != None:
            data = xlrd.open_workbook(self.file_path)
            for sheet in data.sheet_names():
                if sheet_name == sheet:
                    table = data.sheet_by_name(sheet_name=sheet_name)
                    print(table)
                    return table


    # 获取sheet页的列数
    def get_ncols(self):
        if self.table:
            sheet_lines = self.table
            return sheet_lines.nrows
    # 获取sheet页的行数
    def get_norows(self):
        if self.table:
            sheet_lines = self.table
            return sheet_lines.ncols
    # 根据行号获取指定行内容
    # rowx表示是获取第几行的数据
    # start_col表示从索引为多少开始，end_colx表示从索引为多少结束，
    # end_colx为None表示结束没有限制
    # 获取指定行中的数据并以列表的形式返回
    def get_rows(self,row=0):
        if self.table:
            rowx = self.table.row_values(row)
            return rowx
    #根据行号获取指定列的数据
    def get_rowy(self,row=0):
        if self.table:
            rowy = self.table.col_values(row)
            return rowy

    # 获取指定单元格的内容
    def get_cell(self, row, col):
        if self.table:
            cell_data = self.table.cell_value(row, col)
            return cell_data

    # 写入数据
    def write_value(self, row, col,font_color,value=None):
        '''
        写入excel数据
        row : 行
        col : 列
        row,col,value
        '''
        # print(self.file_path)
        # read_data = xlrd.open_workbook(self.file_path)
        # write_data = copy(read_data)
        # sheet_data = write_data.sheet_by_name(sheet_name=self.sheet_name)
        # sheet_data.write(row, col, value)
        # write_data.save(self.file_path)
        workbook = openpyxl.load_workbook(self.file_path)  # 返回一个workbook数据类型的值
        worksheet = workbook[self.sheet_name]  # 获取活动表
        worksheet.cell(row=row, column=col).value = value
        worksheet.cell(row=row, column=col).font = font_color
        # worksheet.write(row, col, value)
        workbook.save(self.file_path)


    # 在一列数据后增加新的列
    def add_column(self,row):
        # # 定义新列
        # data = xlrd.open_workbook(self.file_path)
        # 插入空行空列
        # insert_rows(idx=数字编号, amount=要插入的行数)，插入的行数是在idx行数的下方插入
        # insert_cols(idx=数字编号, amount=要插入的列数)，插入的位置是在idx列数的左侧插入
        workbook = openpyxl.load_workbook(self.file_path)  # 返回一个workbook数据类型的值
        sheet = workbook.active  # 获取活动表
        print('当前活动表是：' + str(sheet))
        if row:
            print(row + 2)
            sheet.insert_cols(idx=row+2, amount=1)
            workbook.save(self.file_path)
            workbook.close()

        # 根据对应的caseid 找到对应行的内容
    def get_rows_data(self, case_id):
        row_num = self.get_row_num(case_id)
        rows_data = self.get_row_values(row_num)
        # print("11111111",rows_data)
        return rows_data

    # 根据对应的caseid找到对应的行号
    def get_row_num(self,case_id):
        print(case_id)
        num = 0
        if self.get_row_data(0):
            clols_data = self.get_row_data(0)
            for col_data in clols_data:
                if case_id in col_data:
                    return num
                num = num + 1

    # 根据行号，找到该行的内容
    def get_row_values(self, row):
        if self.table:
            tables = self.table
            row_data = tables.row_values(row)
            return row_data

    # 根据 行号 获取某一行的内容
    def get_row_data(self,col_id=None):
        if self.table:
            if col_id != None:
                row = self.table.row_values(col_id)
            else:
                row = self.table.row_values(0)
            return row
    # 根据 行号 获取某一列的内容
    def get_num_cols_data(self, col_id=None):
        if self.table:
            if col_id != None:
                cols = self.table.col_values(col_id)
            else:
                cols = self.table.col_values(0)
            # print(cols)
            return cols

    # 根据 匹配文本，获取当前整行的数据
    # 匹配 sales packaging 的表格的数据
    def get_sales_component(self, txt):
        print(txt)
        for row_data in range(self.table.ncols):
            # if re.findall(txt,str(self.table.row_values(row_data))):
            if fuzz.token_set_ratio(txt, str(self.table.col_values(row_data))) == 100:
                # print(self.table.col_values(row_data))
                col_data = self.table.col_values(row_data)
                lst_data = [element for element in col_data if element != ""]
                lst_data = lst_data[lst_data.index('Component'):]


                for li in lst_data:
                    if li == "Component":
                        sales_packaging = self.get_paragraph_between_duplicates(lst_data,"Component")


                        return sales_packaging

    # 匹配 transport packaging 的表格的数据
    def get_transport_component(self, txt):
        print(txt)
        for row_data in range(self.table.ncols):
            # if re.findall(txt,str(self.table.row_values(row_data))):
            if fuzz.token_set_ratio(txt, str(self.table.col_values(row_data))) == 100:
                # print(self.table.col_values(row_data))
                col_data = self.table.col_values(row_data)
                lst_data = [element for element in col_data if element != ""]
                lst_data = lst_data[lst_data.index('Component'):]
                list_data = lst_data[::-1]

                for li in lst_data:
                    if li == "Component":
                        end_index = list_data.index('Component')
                        transport_packaging = list_data[0:end_index]
                        transport_packaging = transport_packaging[::-1]


                        return transport_packaging
    # 截取中间的数据
    def get_paragraph_between_duplicates(self,lst, data):
        start_index = lst.index(data) + 1
        end_index = lst.index(data, start_index)
        return lst[start_index:end_index]

    # 根据 匹配文本，获取当前整行的数据
    def get_cols_data(self, txt,num=0):

        row_list = []
        for row_data in range(self.table.nrows):
            # if re.findall(txt,str(self.table.row_values(row_data))):
            if num == 0:
                # if fuzz.token_set_ratio(txt, str(self.table.row_values(row_data))) == 100:
                # print(str(self.table.row_values(row_data)))
                if self.table.row_values(row_data)[1] == txt:
                    return self.table.row_values(row_data)
            else:
                # if fuzz.token_set_ratio(txt, str(self.table.row_values(row_data))) == 100:
                if self.table.row_values(row_data)[1] == txt:

                    row_list.append(self.table.row_values(row_data))

        return row_list[num]


    # 根据 匹配文本，获取当前文本内容
    def get_text_cols_data(self,txt):
        print(txt)
        for row_data in range(self.table.nrows):
            # if re.findall(txt,str(self.table.row_values(row_data))):
            if fuzz.token_set_ratio(txt,str(self.table.row_values(row_data))) == 100:
                print(self.table.row_values(row_data))
                print(self.table.row_values(row_data)[1])
                return self.table.row_values(row_data)[1]

    # 根据 匹配文本，获取当前文本内容值
    def get_value_cols_data(self,txt):
        print(txt)
        for row_data in range(self.table.nrows):
            if fuzz.token_set_ratio(txt,str(self.table.row_values(row_data))) == 100:
                print(fuzz.token_set_ratio(txt,str(self.table.row_values(row_data))))
                print(self.table.row_values(row_data))
                print(self.table.row_values(row_data)[2])
                return self.table.row_values(row_data)[2]

    # 根据 所有匹配文本，获取当前文本内容值，返回一个列表
    def get_value_cols_list(self,txt):
        print(txt)
        shell_list = []
        for row_data in range(self.table.nrows):
            # print(row_data)
            if fuzz.token_set_ratio(txt,str(self.table.row_values(row_data))) == 100:
                print(self.table.row_values(row_data))

                shell_list.append(self.table.row_values(row_data)[2])

        print(shell_list)
        return shell_list



if __name__ == "__main__":
    operate = OperateExcel('../../data/PERD EOS Gen2 combo.xlsx','Section 5')
    # # operate.get_sheet("Section 1")
    # # print(sheet)
    rl = operate.get_value_cols_data('Enter Product total PCC recycled plastic percentage') * 100
    print(rl)
    # # if rl:
    # #     print("dddad")
    # rows = operate.get_value_cols_list("Use the link above to access the BOM Scrub Verification File")
    # print(rows)
    # # if rows:
    #     print("有值")
    # open_excel = OperateExcel('../../data/Packaging collection-Lenovo KM203W(选件)-20231116.xlsx', '1200x800mm栈板')
    #
    #
    #
    # packaging_sales_list = open_excel.get_sales_component("Component")
    #
    # packaging_Transport_list =  open_excel.get_transport_component("Component")
    # print(packaging_Transport_list)
    # # cc = open_excel.get_cols_data(packaging_Transport_list[0],1)
    # # print(cc)
    # get_row_data = []
    # for i in range(len(packaging_Transport_list)):
    #     for li in packaging_sales_list:
    #         # print("li::",li)
    #         if packaging_Transport_list[i] == li:
    #             get_row_data = open_excel.get_cols_data(packaging_Transport_list[i], 1)
    #             break
    #         else:
    #             get_row_data = open_excel.get_cols_data(packaging_Transport_list[i])
    #     print(get_row_data)
    # rows = open_excel.get_value_cols_list("Width (mm)")
    # print(rows)
    # num = 0
    #
    # #
    # col1 = open_excel.get_value_cols_list("Width(mm)")
    #
    # print(col1)

    # for i in range(len(col1)):
    #     print(col1[i])


    # if "Please list the relevant component PERD reference(s)" in col:
    #     print("进入")
        # print("col1::::",col1)

    # aaa = "Does this product contain part assemblies that have their own PERD documents?"
    # if aaa in col:
    #     print("进了")
    # cols = operate.get_value_cols_data("Attach Power Consumption Test report for external power supply")
    # print(col)
    # print(cols)
    # if "Please attach the required ErP Lot 3 and/or ErP Lot 7 test Report" in col:
    #     print("成功")
    # print(cols[-1])
    # if cols == 'Yes' or cols == 'yes':
    #     print('aa')
    # print(cols[:-2])
    # cols_list = cols.split(',')
    # print(cols_list)
    # for i in cols_list:
    #     print(i)
    # # 行数
    # rows=operate.get_norows()
    # # print(rows)
    # # # 列数
    # coums=operate.get_ncols()
    # print(coums)
    # operate.get_rows(2)
    # rowy_data=operate.get_rowy(3)
    # print(rowy_data[3])
    # cell = operate.get_cell(2,3)
    # num = operate.get_cols_data(1)

    # id = operate.get_row_num('S-Chinese')
    #
    # yan = operate.get_cols_data(id)
    # print(id)
    # operate.add_column(id)
    #
    # lists = operate.get_rows_data("S-Chinese")
    # print(lists)
    # print(cell)





    # lines=operate.get_lines()
    # print(lines)
    # operate.get_cell(1, 1)
    # operate.write_value(1,12,"pass")
